import cv2
import openpyxl
from openpyxl.styles import PatternFill
from tqdm import tqdm
import os

dir_list = os.listdir('./inputs')
outputs_list = os.listdir('./outputs')

if __name__ == '__main__':
    for input in dir_list:
        if input.split('.')[0] in [output.split('.')[0] for output in outputs_list]:
            continue
        image = cv2.imread(f'./inputs/{input}')
        resize_scale = 500 / len(image[0]) if len(image[0]) > len(image) else 500 / len(image)
        resized_image = cv2.resize(image, (int(len(image[0]) * resize_scale), int(len(image)*resize_scale)))

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.sheet_view.showGridLines = False
        sheet.sheet_view.zoomScale = 25
        sheet.sheet_view.showRowColHeaders = False

        bar = tqdm(total=len(resized_image[0])*len(resized_image), desc=f'Processing {input}...: ', position=0)

        for i in range(0, len(resized_image[0])):
            for j in range(0, len(resized_image)):
                sheet.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = 1
                sheet.row_dimensions[j+1].height = 5
                (b, g, r) = resized_image[j, i]
                fill = PatternFill(start_color="FF%02x%02x%02x" % (r, g, b), end_color="FF%02x%02x%02x" % (r, g, b), fill_type="solid")
                sheet.cell(row=j+1, column=i+1).fill = fill
                bar.update(1)

        bar.close()

        # Save the workbook
        workbook.save(f"./outputs/{input.split('.')[0]}.xlsx")